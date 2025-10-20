import openpyxl, os, uuid, datetime, bcrypt
from .repo_interface import Repository

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "../data/board.xlsx")

class ExcelRepository(Repository):
    def __init__(self):
        if not os.path.exists(EXCEL_PATH):
            os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Tickets"
            ws.append(['ticket_id','title','content','author_name','author_contact','pwd_hash','has_admin_reply','status','created_at','updated_at'])
            ws2 = wb.create_sheet('ThreadMessages')
            ws2.append(['msg_id','ticket_id','role','content','created_at'])
            ws3 = wb.create_sheet('AdminUsers')
            ws3.append(['admin_id','username','pwd_hash'])
            # default admin/admin1234
            ws3.append(['0001','admin', bcrypt.hashpw('admin1234'.encode(), bcrypt.gensalt()).decode()])
            ws4 = wb.create_sheet('Attachments')
            ws4.append(['file_id','ticket_id','stored_path','orig_name','mime','size','created_at'])
            wb.save(EXCEL_PATH)

    def _load(self):
        return openpyxl.load_workbook(EXCEL_PATH)

    # Tickets
    def create_ticket(self, ticket):
        wb = self._load(); ws = wb["Tickets"]
        ws.append(ticket); wb.save(EXCEL_PATH)

    def list_tickets(self):
        wb = self._load(); ws = wb["Tickets"]
        return [row for row in ws.iter_rows(values_only=True)][1:]

    def get_ticket(self, ticket_id):
        wb = self._load(); ws = wb["Tickets"]
        for row in ws.iter_rows(values_only=True):
            if row[0] == ticket_id:
                return row
        return None

    def update_ticket(self, ticket_id, data):
        wb = self._load(); ws = wb["Tickets"]
        headers = [c.value for c in ws[1]]
        idx = {h:i for i,h in enumerate(headers)}
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            if r_i == 0: continue
            if row[0] == ticket_id:
                for k,v in data.items():
                    if k in idx:
                        ws.cell(row=r_i+1, column=idx[k]+1, value=v)
                break
        wb.save(EXCEL_PATH)

    def delete_ticket(self, ticket_id):
        wb = self._load(); ws = wb["Tickets"]
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            if r_i == 0: continue
            if row[0] == ticket_id:
                ws.delete_rows(r_i+1, 1); break
        wb.save(EXCEL_PATH)

    # Messages
    def create_message(self, m):
        wb = self._load(); ws = wb["ThreadMessages"]
        ws.append([m['msg_id'], m['ticket_id'], m['role'], m['content'], m['created_at']]); wb.save(EXCEL_PATH)

    def list_messages(self, ticket_id):
        wb = self._load(); ws = wb["ThreadMessages"]
        out = []
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            if r_i == 0: continue
            if row[1] == ticket_id:
                out.append({'msg_id':row[0],'ticket_id':row[1],'role':row[2],'content':row[3],'created_at':row[4]})
        return out

    def mark_has_admin_reply(self, ticket_id):
        wb = self._load(); ws = wb["Tickets"]
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            if r_i == 0: continue
            if row[0] == ticket_id:
                ws.cell(row=r_i+1, column=7, value=1)  # has_admin_reply
                ws.cell(row=r_i+1, column=8, value='ANSWERED')
                break
        wb.save(EXCEL_PATH)

    # Admin
    def get_admin_user(self, username):
        wb = self._load(); ws = wb["AdminUsers"]
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            if r_i == 0: continue
            if row[1] == username:
                return {'admin_id':row[0],'username':row[1],'pwd_hash':row[2]}
        return None

    # Attachments
    def create_attachment(self, ticket_id, stored_path, orig_name, mime, size):
        wb = self._load(); ws = wb["Attachments"]
        ws.append([str(uuid.uuid4()), ticket_id, stored_path, orig_name, mime, size, datetime.datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
